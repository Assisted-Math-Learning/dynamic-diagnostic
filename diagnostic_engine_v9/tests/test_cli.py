"""Tests for the CLI subcommands (seed-config, seed-lattice, simulate-session, validate-config)."""

import io
from pathlib import Path

from tests import DATA_DIR
from typing import List

import openpyxl
import pytest
import yaml

from engine.cli import main
from engine.config import load_engine_config

TEST_FIXTURE = Path(__file__).parent / "fixtures" / "engine_config_test.yaml"


# Helpers --------------------------------------------------------------------


def _capture(argv: List[str]) -> tuple:
    """Run main(argv) and capture stderr; return (exit_code, stderr_text)."""
    buf = io.StringIO()
    code = main(argv, stderr=buf)
    return code, buf.getvalue()


def _make_synthetic_inputs(tmp_path: Path) -> dict:
    """Create a small set of synthetic input files mirroring the real shapes."""
    ms = tmp_path / "milestone.csv"
    ms.write_text(
        "L2_5_sequence,qset_class,L1_skill,L2_5_skill,qset_purpose,count_of_total_diagnostic_qsets\n"
        "1,class-one,Addition,1D+1D sum upto 9,Micro Diagnostic,1\n"
        "2,class-two,Multiplication,Repeated addition,Micro Diagnostic,1\n"
        "3,class-three,Multiplication,Tables 1 to 9,Micro Diagnostic,1\n"
        "4,class-one,Subtraction,1D - 0 to 9,Micro Diagnostic,1\n"
        "5,class-two,Division,Division using Distribution,Micro Diagnostic,1\n"
        "6,class-three,Addition,2-digit Addition with carry,Micro Diagnostic,1\n",
        encoding="utf-8",
    )

    priors = tmp_path / "priors.csv"
    priors.write_text(
        "Student Class,operation,L2_5_sequence,skill_name,n,n_mastered,p_mastered,mean_score,thin\n"
        "3,Addition,1,1D+1D sum upto 9,100,95,0.95,0.9,False\n"
        "3,Addition,6,2-digit Addition with carry,100,60,0.6,0.65,False\n"
        "3,Multiplication,2,Repeated addition,100,85,0.85,0.8,False\n"
        "3,Multiplication,3,Tables 1 to 9,100,50,0.5,0.6,False\n"
        "3,Subtraction,4,1D - 0 to 9,100,95,0.95,0.9,False\n"
        "3,Division,5,Division using Distribution,100,50,0.5,0.6,False\n"
        "2,Addition,1,1D+1D sum upto 9,100,85,0.85,0.85,False\n"
        "2,Multiplication,2,Repeated addition,100,50,0.5,0.6,False\n"
        "2,Subtraction,4,1D - 0 to 9,100,85,0.85,0.85,False\n"
        "2,Division,5,Division using Distribution,100,30,0.3,0.4,False\n",
        encoding="utf-8",
    )

    anchors = tmp_path / "anchors.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "README"
    ws = wb.create_sheet("Recommended anchor per grade")
    ws.append(["Learner grade", "Operation", "Recommended anchor skill"])
    for row in [
        (2, "Addition", "1D+1D sum upto 9"),
        (2, "Subtraction", "1D - 0 to 9"),
        (2, "Multiplication", "Repeated addition"),
        (2, "Division", "Division using Distribution"),
        (3, "Addition", "2-digit Addition with carry"),
        (3, "Subtraction", "1D - 0 to 9"),
        (3, "Multiplication", "Tables 1 to 9"),
        (3, "Division", "Division using Distribution"),
        (4, "Addition", "2-digit Addition with carry"),
        (4, "Subtraction", "1D - 0 to 9"),
        (4, "Multiplication", "Tables 1 to 9"),
        (4, "Division", "Division using Distribution"),
        (5, "Addition", "2-digit Addition with carry"),
        (5, "Subtraction", "1D - 0 to 9"),
        (5, "Multiplication", "Tables 1 to 9"),
        (5, "Division", "Division using Distribution"),
    ]:
        ws.append(list(row))
    wb.save(anchors)
    wb.close()

    lat = tmp_path / "lattice.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Operation A", "Skill A", "Content grade of Skill A",
        "Operation B", "Skill B", "Content grade of Skill B",
        "Within the same operation?", "Skill B lower than Skill A in AML content flow?",
        "# measurements where edge passes", "Measurements where edge passes",
        "Telangana Exit→Entry: P(B|A) — pooled",
        "Telangana Exit→Entry: P(B|not A) — pooled",
        "Delhi Entry→Entry: P(B|A) — pooled",
        "Delhi Entry→Entry: P(B|not A) — pooled",
    ]
    ws.append(headers)
    ws.append([
        "Addition", "2-digit Addition with carry", "class-three",
        "Addition", "1D+1D sum upto 9", "class-one",
        "Y", "Yes",
        2, "DL + TG",
        0.92, 0.55, 0.94, 0.60,
    ])
    wb.save(lat)
    wb.close()

    return {
        "milestone": str(ms),
        "priors": str(priors),
        "anchors": str(anchors),
        "lattice": str(lat),
    }


# === seed-config ============================================================


class TestSeedConfig:
    def test_happy_path(self, tmp_path: Path):
        inputs = _make_synthetic_inputs(tmp_path)
        output = tmp_path / "engine_config.yaml"
        code, stderr = _capture([
            "seed-config",
            "--milestone-mapping", inputs["milestone"],
            "--priors", inputs["priors"],
            "--anchors", inputs["anchors"],
            "--output", str(output),
        ])
        assert code == 0, stderr
        assert output.exists()

    def test_output_yaml_roundtrips(self, tmp_path: Path):
        inputs = _make_synthetic_inputs(tmp_path)
        output = tmp_path / "engine_config.yaml"
        code, _ = _capture([
            "seed-config",
            "--milestone-mapping", inputs["milestone"],
            "--priors", inputs["priors"],
            "--anchors", inputs["anchors"],
            "--output", str(output),
        ])
        assert code == 0
        # The output must load through load_engine_config without error
        config = load_engine_config(str(output))
        assert config.version == "0.7.0"
        # G3 anchor for Mult should be "Tables 1 to 9" per synthetic anchors
        assert config.anchors[3]["Multiplication"] == "Tables 1 to 9"
        # G3 prior for that skill should be 0.5
        assert config.priors[3]["Tables 1 to 9"] == 0.5

    def test_grades_without_priors_get_empty_dict(self, tmp_path: Path):
        """Match the real-data case where G2 has anchors but no priors."""
        inputs = _make_synthetic_inputs(tmp_path)
        # Strip G2 rows from priors CSV
        priors_text = Path(inputs["priors"]).read_text()
        filtered = "\n".join(
            line for line in priors_text.splitlines()
            if not line.startswith("2,")
        )
        Path(inputs["priors"]).write_text(filtered)

        output = tmp_path / "engine_config.yaml"
        code, _ = _capture([
            "seed-config",
            "--milestone-mapping", inputs["milestone"],
            "--priors", inputs["priors"],
            "--anchors", inputs["anchors"],
            "--output", str(output),
        ])
        assert code == 0
        config = load_engine_config(str(output))
        assert 2 in config.priors
        assert config.priors[2] == {}

    def test_missing_input_file_returns_nonzero(self, tmp_path: Path):
        inputs = _make_synthetic_inputs(tmp_path)
        code, stderr = _capture([
            "seed-config",
            "--milestone-mapping", "/nonexistent.csv",
            "--priors", inputs["priors"],
            "--anchors", inputs["anchors"],
            "--output", str(tmp_path / "out.yaml"),
        ])
        assert code != 0
        assert "ERROR" in stderr

    def test_against_real_project_data(self, tmp_path: Path):
        """End-to-end against the real CSVs/XLSXs in the repo data dir."""
        files = {
            "milestone": str(DATA_DIR / "20260518_AML_Telangana_Milestone_and_Level_Mapping.csv"),
            "priors": str(DATA_DIR / "priors_table_delhi_only.csv"),
            "anchors": str(DATA_DIR / "anchor_recommendations_v3.xlsx"),
        }
        for f in files.values():
            if not Path(f).exists():
                pytest.skip("real data files not available")

        output = tmp_path / "engine_config.yaml"
        code, _ = _capture([
            "seed-config",
            "--milestone-mapping", files["milestone"],
            "--priors", files["priors"],
            "--anchors", files["anchors"],
            "--output", str(output),
        ])
        assert code == 0
        config = load_engine_config(str(output))
        # 39 canonical L2.5 skills across 4 ops
        assert len(config.skills) == 39
        assert sorted(config.budgets) == [2, 3, 4, 5]


# === seed-lattice ===========================================================


class TestSeedLattice:
    def test_happy_path(self, tmp_path: Path):
        inputs = _make_synthetic_inputs(tmp_path)
        code, stderr = _capture([
            "seed-lattice",
            "--input", inputs["lattice"],
        ])
        assert code == 0
        assert "loaded 1 edges" in stderr

    def test_against_real_project_data(self):
        path = str(DATA_DIR / "lattice_edges_final.xlsx")
        if not Path(path).exists():
            pytest.skip("real lattice file not available")
        code, stderr = _capture(["seed-lattice", "--input", path])
        assert code == 0
        assert "loaded 12 edges" in stderr


# === simulate-session ======================================================


class TestSimulateSession:
    def test_all_correct_completes(self, tmp_path: Path):
        """G3 simulate-session with the test fixture YAML and all-correct policy."""
        code, stderr = _capture([
            "simulate-session",
            "--grade", "3",
            "--policy", "all-correct",
            "--config", str(TEST_FIXTURE),
        ])
        assert code == 0
        assert "summary" in stderr
        assert "end reason" in stderr
        assert "questions asked" in stderr

    def test_all_incorrect_completes(self, tmp_path: Path):
        code, stderr = _capture([
            "simulate-session",
            "--grade", "3",
            "--policy", "all-incorrect",
            "--config", str(TEST_FIXTURE),
        ])
        assert code == 0
        assert "summary" in stderr

    def test_unconfigured_grade_returns_nonzero(self):
        """Test fixture only has G2/G3; G4 is not configured."""
        code, stderr = _capture([
            "simulate-session",
            "--grade", "4",
            "--policy", "all-correct",
            "--config", str(TEST_FIXTURE),
        ])
        assert code != 0
        assert "ERROR" in stderr

    def test_with_real_data_g3(self, tmp_path: Path):
        """End-to-end: seed-config from real CSVs, then simulate G3."""
        files = {
            "milestone": str(DATA_DIR / "20260518_AML_Telangana_Milestone_and_Level_Mapping.csv"),
            "priors": str(DATA_DIR / "priors_table_delhi_only.csv"),
            "anchors": str(DATA_DIR / "anchor_recommendations_v3.xlsx"),
            "lattice": str(DATA_DIR / "lattice_edges_final.xlsx"),
        }
        for f in files.values():
            if not Path(f).exists():
                pytest.skip("real data files not available")

        config_path = tmp_path / "engine_config.yaml"
        code, _ = _capture([
            "seed-config",
            "--milestone-mapping", files["milestone"],
            "--priors", files["priors"],
            "--anchors", files["anchors"],
            "--output", str(config_path),
        ])
        assert code == 0

        code, stderr = _capture([
            "simulate-session",
            "--grade", "3",
            "--policy", "all-correct",
            "--config", str(config_path),
            "--lattice", files["lattice"],
        ])
        assert code == 0
        assert "questions asked" in stderr
        # G3 has 21 skills in scope; the engine must produce verdicts for all
        assert "verdicts        : 21 skills" in stderr


# === validate-config ========================================================


class TestValidateConfig:
    def test_valid_config_returns_zero(self):
        code, stderr = _capture(["validate-config", "--config", str(TEST_FIXTURE)])
        assert code == 0
        assert "valid" in stderr
        assert "G2: builds OK" in stderr
        assert "G3: builds OK" in stderr

    def test_missing_file_returns_nonzero(self):
        code, stderr = _capture(["validate-config", "--config", "/nonexistent.yaml"])
        assert code != 0
        assert "ERROR" in stderr

    def test_malformed_yaml_returns_nonzero(self, tmp_path: Path):
        bad = tmp_path / "bad.yaml"
        bad.write_text("this: is: not: valid: yaml: [\n", encoding="utf-8")
        code, stderr = _capture(["validate-config", "--config", str(bad)])
        assert code != 0

    def test_invalid_config_structure_returns_nonzero(self, tmp_path: Path):
        bad = tmp_path / "bad.yaml"
        bad.write_text("version: 0.1.0\nalgorithm: {}", encoding="utf-8")  # missing fields
        code, stderr = _capture(["validate-config", "--config", str(bad)])
        assert code != 0
        assert "ERROR" in stderr


class TestCleanup:
    """`python -m engine.cli cleanup` end-to-end via main() (fix-pack change #5).

    The cleanup CLI itself uses get_storage_backend() which defaults to
    InMemoryStorage when STORAGE_BACKEND is unset. An InMemoryStorage
    instance is fresh per CLI process, so the CLI run finds zero sessions
    (the typical real-deployment case is MongoDB-backed and shared across
    runs). These tests cover: config loading errors, success path, and
    storage selection.
    """

    def test_missing_config_returns_2(self, tmp_path):
        code, stderr = _capture([
            "cleanup", "--config", str(tmp_path / "does_not_exist.yaml"),
        ])
        assert code == 2
        assert "cannot load config" in stderr

    def test_invalid_config_returns_2(self, tmp_path):
        bad = tmp_path / "bad.yaml"
        bad.write_text("version: 0.1.0\nalgorithm: {}", encoding="utf-8")
        code, stderr = _capture(["cleanup", "--config", str(bad)])
        assert code == 2
        assert "cannot load config" in stderr

    def test_success_with_no_sessions_in_storage_returns_0(self, tmp_path, monkeypatch):
        """Default InMemoryStorage with no sessions -> cleanup exits 0 (no work to do)."""
        # Ensure storage is memory regardless of env
        monkeypatch.setenv("STORAGE_BACKEND", "memory")
        code, _ = _capture(["cleanup", "--config", str(TEST_FIXTURE)])
        assert code == 0

    def test_explicit_limit_flag_accepted(self, tmp_path, monkeypatch):
        """--limit is parsed and accepted."""
        monkeypatch.setenv("STORAGE_BACKEND", "memory")
        code, _ = _capture([
            "cleanup", "--config", str(TEST_FIXTURE), "--limit", "50",
        ])
        assert code == 0

    def test_explicit_storage_flag_accepted(self, tmp_path, monkeypatch):
        """--storage memory is parsed and accepted."""
        code, _ = _capture([
            "cleanup", "--config", str(TEST_FIXTURE), "--storage", "memory",
        ])
        assert code == 0

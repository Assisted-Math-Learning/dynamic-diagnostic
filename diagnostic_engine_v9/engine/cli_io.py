"""
Pure-Python loaders for the canonical AML data files used by the CLI.

These functions take file paths and return plain dicts / lists. They DO NOT
construct EngineConfig or LatticeIndex objects directly - the CLI subcommands
do that. Keeping the loaders separate makes them easy to unit-test against
small synthetic inputs.

The functions are deliberately strict: they raise CliInputError (with a
descriptive message) when a required column is missing, a value is malformed,
or a cross-reference is broken. The CLI's main() catches CliInputError and
prints a clean error rather than a Python traceback.
"""

import csv
from pathlib import Path
from typing import Dict, List, Set, Tuple

import openpyxl

from engine.lattice import LatticeEdge

# Map qset_class string ('class-one', 'class-two', ...) to integer content grade.
_QSET_CLASS_TO_GRADE: Dict[str, int] = {
    "class-one": 1,
    "class-two": 2,
    "class-three": 3,
    "class-four": 4,
    "class-five": 5,
    "class-six": 6,
    "class-seven": 7,
    "class-eight": 8,
}


class CliInputError(ValueError):
    """Raised by the loaders when an input file is malformed or inconsistent."""


# === Milestone mapping CSV =================================================


def load_milestone_mapping(
    path: str, *, allowed_operations: Set[str],
) -> List[Dict]:
    """Load the canonical L2.5 skill list from the milestone mapping CSV.

    Required columns: L2_5_sequence, qset_class, L1_skill, L2_5_skill.
    Returns a list of dicts with keys: name, operation, sequence, content_grade.

    Skills whose operation is not in `allowed_operations` are skipped (this is
    how Numbers is excluded from the engine's scope).
    """
    p = Path(path)
    if not p.exists():
        raise CliInputError(f"milestone mapping not found: {path}")

    required = {"L2_5_sequence", "qset_class", "L1_skill", "L2_5_skill"}
    skills: List[Dict] = []

    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise CliInputError(f"milestone mapping {path} has no header row")
        missing = required - set(reader.fieldnames)
        if missing:
            raise CliInputError(
                f"milestone mapping {path} missing columns: {sorted(missing)}"
            )

        for line_no, row in enumerate(reader, start=2):  # header is line 1
            operation = (row["L1_skill"] or "").strip()
            if operation not in allowed_operations:
                continue
            name = (row["L2_5_skill"] or "").strip()
            if not name:
                raise CliInputError(
                    f"{path}:{line_no} L2_5_skill is empty"
                )
            try:
                sequence = int(row["L2_5_sequence"])
            except (TypeError, ValueError):
                raise CliInputError(
                    f"{path}:{line_no} L2_5_sequence '{row['L2_5_sequence']}' is not an integer"
                )
            qset_class = (row["qset_class"] or "").strip()
            if qset_class not in _QSET_CLASS_TO_GRADE:
                raise CliInputError(
                    f"{path}:{line_no} qset_class '{qset_class}' unrecognised; "
                    f"expected one of {sorted(_QSET_CLASS_TO_GRADE)}"
                )
            skills.append({
                "name": name,
                "operation": operation,
                "sequence": sequence,
                "content_grade": _QSET_CLASS_TO_GRADE[qset_class],
            })

    # Deduplicate by skill name (taking the first occurrence; the CSV might list
    # the same L2.5 skill across multiple qset_purpose rows).
    seen: Set[str] = set()
    deduped: List[Dict] = []
    for s in skills:
        if s["name"] in seen:
            continue
        seen.add(s["name"])
        deduped.append(s)
    return deduped


# === Priors CSV ============================================================


def load_priors(
    path: str, *, allowed_operations: Set[str],
) -> Dict[int, Dict[str, float]]:
    """Load per-(grade, skill) cohort priors from a priors CSV.

    The canonical input is `priors_table_delhi_only.csv`, derived from the
    raw Delhi diagnostic response population (n_DL >= 130 per skill). An
    earlier `priors_table.csv` was computed from MainD response data
    instead of the raw diagnostic and produced uniformly-high, biased
    priors; do not use it. Either filename loads through this function as
    long as the column shape is right.

    Required columns: Student Class, operation, skill_name, p_mastered.
    Returns {grade: {skill_name: prior}}. Skips rows whose operation is not in
    allowed_operations.
    """
    p = Path(path)
    if not p.exists():
        raise CliInputError(f"priors table not found: {path}")

    required = {"Student Class", "operation", "skill_name", "p_mastered"}
    priors: Dict[int, Dict[str, float]] = {}

    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise CliInputError(f"priors table {path} has no header row")
        missing = required - set(reader.fieldnames)
        if missing:
            raise CliInputError(
                f"priors table {path} missing columns: {sorted(missing)}"
            )

        for line_no, row in enumerate(reader, start=2):
            operation = (row["operation"] or "").strip()
            if operation not in allowed_operations:
                continue
            try:
                grade = int(row["Student Class"])
            except (TypeError, ValueError):
                raise CliInputError(
                    f"{path}:{line_no} Student Class '{row['Student Class']}' is not an integer"
                )
            skill = (row["skill_name"] or "").strip()
            if not skill:
                raise CliInputError(f"{path}:{line_no} skill_name is empty")
            try:
                prior = float(row["p_mastered"])
            except (TypeError, ValueError):
                raise CliInputError(
                    f"{path}:{line_no} p_mastered '{row['p_mastered']}' is not a float"
                )
            if not 0.0 <= prior <= 1.0:
                raise CliInputError(
                    f"{path}:{line_no} p_mastered {prior} out of range [0, 1]"
                )
            priors.setdefault(grade, {})[skill] = prior
    return priors


# === Anchor recommendations XLSX ===========================================


_ANCHOR_SHEET = "Recommended anchor per grade"
_ANCHOR_REQUIRED_COLS = {"Learner grade", "Operation", "Recommended anchor skill"}


def load_anchors(
    path: str, *, allowed_operations: Set[str],
) -> Dict[int, Dict[str, str]]:
    """Load per-(grade, operation) anchor skill from anchor_recommendations_v3.xlsx.

    Reads the sheet 'Recommended anchor per grade'. Required columns:
    'Learner grade', 'Operation', 'Recommended anchor skill'.

    Returns {grade: {operation: anchor_skill}}.
    """
    p = Path(path)
    if not p.exists():
        raise CliInputError(f"anchors file not found: {path}")

    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        if _ANCHOR_SHEET not in wb.sheetnames:
            raise CliInputError(
                f"anchors file {path} missing sheet '{_ANCHOR_SHEET}'; "
                f"sheets present: {wb.sheetnames}"
            )
        ws = wb[_ANCHOR_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise CliInputError(f"anchors sheet '{_ANCHOR_SHEET}' is empty")
        header_index = {name: i for i, name in enumerate(header) if name is not None}
        missing = _ANCHOR_REQUIRED_COLS - set(header_index)
        if missing:
            raise CliInputError(
                f"anchors sheet '{_ANCHOR_SHEET}' missing columns: {sorted(missing)}"
            )

        anchors: Dict[int, Dict[str, str]] = {}
        for line_no, row in enumerate(rows, start=2):
            grade_raw = row[header_index["Learner grade"]]
            if grade_raw is None:
                continue  # tolerate trailing blank rows
            try:
                grade = int(grade_raw)
            except (TypeError, ValueError):
                raise CliInputError(
                    f"{path}:{_ANCHOR_SHEET}:{line_no} Learner grade '{grade_raw}' is not an integer"
                )
            operation = (row[header_index["Operation"]] or "").strip()
            if operation not in allowed_operations:
                continue
            skill = (row[header_index["Recommended anchor skill"]] or "").strip()
            if not skill:
                raise CliInputError(
                    f"{path}:{_ANCHOR_SHEET}:{line_no} anchor skill is empty for grade {grade} {operation}"
                )
            anchors.setdefault(grade, {})[operation] = skill
        return anchors
    finally:
        wb.close()


# === Lattice edges XLSX ====================================================


_LATTICE_REQUIRED_COLS = {
    "Operation A", "Skill A", "Operation B", "Skill B",
    "# measurements where edge passes",
    "Delhi Entry→Entry: P(B|A) — pooled",
    "Delhi Entry→Entry: P(B|not A) — pooled",
}


def load_lattice_edges(path: str) -> List[LatticeEdge]:
    """Load lattice edges from lattice_edges_final.xlsx.

    Reads the first sheet. Required columns: Operation A/B, Skill A/B,
    '# measurements where edge passes', and the Delhi Entry→Entry P(B|A) and
    P(B|not A) columns.

    Edge values:
      - p_b_given_a, p_b_given_not_a: read from Delhi Entry→Entry columns
        ONLY. The engine was calibrated on Delhi data, so edge strengths
        must come from the same source. The Telangana Exit→Entry columns
        measure a different signal (mastery persistence across the class
        boundary, not co-mastery at a single point in time) on a different
        cohort and are not interchangeable. Switching the engine to use
        Telangana edge values is a tracked workstream and is out of scope
        for this loader.
      - weight: '# measurements where edge passes' >= 2 -> multi-view (1.0);
        single-view -> 0.5.

    Returns a list of LatticeEdge objects.
    """
    p = Path(path)
    if not p.exists():
        raise CliInputError(f"lattice edges file not found: {path}")

    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise CliInputError(f"lattice file {path} is empty")
        header_index = {name: i for i, name in enumerate(header) if name is not None}
        missing = _LATTICE_REQUIRED_COLS - set(header_index)
        if missing:
            raise CliInputError(
                f"lattice file {path} missing columns: {sorted(missing)}"
            )

        edges: List[LatticeEdge] = []
        for line_no, row in enumerate(rows, start=2):
            skill_a_raw = row[header_index["Skill A"]]
            if skill_a_raw is None:
                continue  # trailing blank rows
            skill_a = str(skill_a_raw).strip()
            skill_b = str(row[header_index["Skill B"]] or "").strip()
            operation_a = str(row[header_index["Operation A"]] or "").strip()
            operation_b = str(row[header_index["Operation B"]] or "").strip()
            if not (skill_a and skill_b and operation_a and operation_b):
                raise CliInputError(
                    f"{path}:{line_no} skill_a/b or operation_a/b is empty"
                )

            p_b_given_a = row[header_index["Delhi Entry→Entry: P(B|A) — pooled"]]
            p_b_given_not_a = row[header_index["Delhi Entry→Entry: P(B|not A) — pooled"]]
            if p_b_given_a is None or p_b_given_not_a is None:
                raise CliInputError(
                    f"{path}:{line_no} edge {skill_a} -> {skill_b}: "
                    "Delhi Entry→Entry P(B|A) or P(B|not A) is missing. "
                    "The engine requires Delhi edge values since calibration is Delhi-based."
                )

            try:
                p_b_given_a = float(p_b_given_a)
                p_b_given_not_a = float(p_b_given_not_a)
                n_measurements = int(row[header_index["# measurements where edge passes"]])
            except (TypeError, ValueError) as e:
                raise CliInputError(
                    f"{path}:{line_no} numeric column not parsable: {e}"
                )
            weight = 1.0 if n_measurements >= 2 else 0.5
            edges.append(LatticeEdge(
                skill_a=skill_a,
                skill_b=skill_b,
                operation_a=operation_a,
                operation_b=operation_b,
                p_b_given_a=p_b_given_a,
                p_b_given_not_a=p_b_given_not_a,
                weight=weight,
            ))
        return edges
    finally:
        wb.close()


# === Cross-validation helpers ==============================================


def cross_validate(
    *,
    skills: List[Dict],
    anchors: Dict[int, Dict[str, str]],
    priors: Dict[int, Dict[str, float]],
) -> List[str]:
    """Check internal consistency of the loaded files. Returns a list of warnings.

    Empty list means clean. Each warning is a human-readable string. The CLI's
    seed-config prints these to stderr but proceeds (warnings, not errors) so
    that small data gaps don't block the YAML generation.
    """
    warnings: List[str] = []
    skill_names = {s["name"] for s in skills}

    for grade, ops in anchors.items():
        for op, anchor_skill in ops.items():
            if anchor_skill not in skill_names:
                warnings.append(
                    f"anchor for G{grade} {op} is '{anchor_skill}' but that skill "
                    "is not in the milestone mapping"
                )

    for grade, prior_map in priors.items():
        for skill in prior_map:
            if skill not in skill_names:
                warnings.append(
                    f"prior for G{grade} skill '{skill}' but that skill is not "
                    "in the milestone mapping"
                )

    return warnings

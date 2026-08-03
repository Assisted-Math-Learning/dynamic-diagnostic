"""Unit tests for the CLI loaders in engine.cli_io."""

from pathlib import Path

from tests import DATA_DIR
from typing import List

import openpyxl
import pytest

from engine.cli_io import (
    CliInputError,
    cross_validate,
    load_anchors,
    load_lattice_edges,
    load_milestone_mapping,
    load_priors,
)

ALL_OPS = {"Addition", "Subtraction", "Multiplication", "Division"}


# Synthetic file builders -----------------------------------------------------


def _write_csv(path: Path, header: List[str], rows: List[List[str]]) -> None:
    lines = [",".join(header)]
    for r in rows:
        lines.append(",".join(r))
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_anchors_xlsx(
    path: Path,
    rows: List[tuple],
    *,
    sheet_name: str = "Recommended anchor per grade",
    include_extra_cols: bool = True,
) -> None:
    wb = openpyxl.Workbook()
    # Default sheet for unrelated content
    wb.active.title = "README"
    ws = wb.create_sheet(sheet_name)
    headers = ["Learner grade", "Operation", "Recommended anchor skill"]
    if include_extra_cols:
        headers += ["Content grade of anchor", "Anchor selection basis"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    wb.close()


def _write_lattice_xlsx(path: Path, rows: List[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Operation A", "Skill A", "Content grade of Skill A",
        "Operation B", "Skill B", "Content grade of Skill B",
        "Within the same operation?", "Skill B lower than Skill A in AML content flow?",
        "# measurements where edge passes", "Measurements where edge passes",
        "Telangana Exit→Entry: P(B|A) — pooled",
        "Telangana Exit→Entry: P(B|not A) — pooled",
        "Delhi Entry→Entry: P(B|A) — pooled",
        "Delhi Entry→Entry: P(B|not A) — pooled",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    wb.save(path)
    wb.close()


# === load_milestone_mapping =================================================


class TestLoadMilestoneMapping:
    def test_happy_path(self, tmp_path: Path):
        path = tmp_path / "ms.csv"
        _write_csv(
            path,
            ["L2_5_sequence", "qset_class", "L1_skill", "L2_5_skill",
             "qset_purpose", "count_of_total_diagnostic_qsets"],
            [
                ["1", "class-one", "Addition", "1D+1D sum upto 9", "Micro Diagnostic", "1"],
                ["2", "class-two", "Multiplication", "Repeated addition", "Micro Diagnostic", "1"],
                ["3", "class-three", "Numbers", "Place value chart", "Micro Diagnostic", "1"],
            ],
        )
        skills = load_milestone_mapping(str(path), allowed_operations=ALL_OPS)
        # Numbers row is excluded (not in ALL_OPS)
        assert len(skills) == 2
        assert skills[0]["name"] == "1D+1D sum upto 9"
        assert skills[0]["operation"] == "Addition"
        assert skills[0]["sequence"] == 1
        assert skills[0]["content_grade"] == 1
        assert skills[1]["content_grade"] == 2

    def test_dedupes_by_name(self, tmp_path: Path):
        path = tmp_path / "ms.csv"
        _write_csv(
            path,
            ["L2_5_sequence", "qset_class", "L1_skill", "L2_5_skill",
             "qset_purpose", "count_of_total_diagnostic_qsets"],
            [
                ["1", "class-one", "Addition", "1D+1D sum upto 9", "Micro Diagnostic", "1"],
                ["1", "class-one", "Addition", "1D+1D sum upto 9", "Exit Main Diagnostic", "1"],
            ],
        )
        skills = load_milestone_mapping(str(path), allowed_operations=ALL_OPS)
        assert len(skills) == 1

    def test_missing_file_raises(self):
        with pytest.raises(CliInputError, match="not found"):
            load_milestone_mapping("/nonexistent.csv", allowed_operations=ALL_OPS)

    def test_missing_column_raises(self, tmp_path: Path):
        path = tmp_path / "ms.csv"
        _write_csv(path, ["L2_5_sequence", "qset_class"], [["1", "class-one"]])
        with pytest.raises(CliInputError, match="missing columns"):
            load_milestone_mapping(str(path), allowed_operations=ALL_OPS)

    def test_bad_qset_class_raises(self, tmp_path: Path):
        path = tmp_path / "ms.csv"
        _write_csv(
            path,
            ["L2_5_sequence", "qset_class", "L1_skill", "L2_5_skill",
             "qset_purpose", "count_of_total_diagnostic_qsets"],
            [["1", "kindergarten", "Addition", "x", "Micro Diagnostic", "1"]],
        )
        with pytest.raises(CliInputError, match="qset_class"):
            load_milestone_mapping(str(path), allowed_operations=ALL_OPS)


# === load_priors ============================================================


class TestLoadPriors:
    def test_happy_path(self, tmp_path: Path):
        path = tmp_path / "p.csv"
        _write_csv(
            path,
            ["Student Class", "operation", "L2_5_sequence", "skill_name",
             "n", "n_mastered", "p_mastered", "mean_score", "thin"],
            [
                ["3", "Addition", "1", "1D+1D sum upto 9", "100", "85", "0.85", "0.8", "False"],
                ["3", "Numbers", "5", "Place value", "100", "60", "0.6", "0.7", "False"],
                ["4", "Multiplication", "10", "Tables 1 to 5", "200", "150", "0.75", "0.7", "False"],
            ],
        )
        priors = load_priors(str(path), allowed_operations=ALL_OPS)
        # Numbers row is excluded
        assert priors == {
            3: {"1D+1D sum upto 9": 0.85},
            4: {"Tables 1 to 5": 0.75},
        }

    def test_out_of_range_prior_raises(self, tmp_path: Path):
        path = tmp_path / "p.csv"
        _write_csv(
            path,
            ["Student Class", "operation", "L2_5_sequence", "skill_name",
             "n", "n_mastered", "p_mastered", "mean_score", "thin"],
            [["3", "Addition", "1", "x", "1", "1", "1.5", "0.8", "False"]],
        )
        with pytest.raises(CliInputError, match="out of range"):
            load_priors(str(path), allowed_operations=ALL_OPS)

    def test_non_numeric_grade_raises(self, tmp_path: Path):
        path = tmp_path / "p.csv"
        _write_csv(
            path,
            ["Student Class", "operation", "L2_5_sequence", "skill_name",
             "n", "n_mastered", "p_mastered", "mean_score", "thin"],
            [["three", "Addition", "1", "x", "1", "1", "0.5", "0.8", "False"]],
        )
        with pytest.raises(CliInputError, match="Student Class"):
            load_priors(str(path), allowed_operations=ALL_OPS)


# === load_anchors ===========================================================


class TestLoadAnchors:
    def test_happy_path(self, tmp_path: Path):
        path = tmp_path / "a.xlsx"
        _write_anchors_xlsx(
            path,
            rows=[
                (2, "Addition", "1D+1D sum upto 9"),
                (2, "Multiplication", "Repeated addition"),
                (3, "Addition", "2-digit Addition with carry"),
                (2, "Numbers", "Place value chart"),  # should be excluded
            ],
        )
        anchors = load_anchors(str(path), allowed_operations=ALL_OPS)
        assert anchors == {
            2: {"Addition": "1D+1D sum upto 9", "Multiplication": "Repeated addition"},
            3: {"Addition": "2-digit Addition with carry"},
        }

    def test_missing_sheet_raises(self, tmp_path: Path):
        path = tmp_path / "a.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(path)
        with pytest.raises(CliInputError, match="missing sheet"):
            load_anchors(str(path), allowed_operations=ALL_OPS)

    def test_missing_required_column_raises(self, tmp_path: Path):
        path = tmp_path / "a.xlsx"
        # Build XLSX with the correct sheet name but missing the required column
        wb = openpyxl.Workbook()
        wb.active.title = "Other"
        ws = wb.create_sheet("Recommended anchor per grade")
        ws.append(["Learner grade", "Operation"])  # missing "Recommended anchor skill"
        ws.append([2, "Addition"])
        wb.save(path)
        with pytest.raises(CliInputError, match="missing columns"):
            load_anchors(str(path), allowed_operations=ALL_OPS)


# === load_lattice_edges =====================================================


class TestLoadLatticeEdges:
    def test_happy_path_delhi_data(self, tmp_path: Path):
        """Edge values come from Delhi Entry→Entry columns."""
        path = tmp_path / "lat.xlsx"
        _write_lattice_xlsx(path, rows=[
            {
                "Operation A": "Addition", "Skill A": "2-digit Addition with carry",
                "Operation B": "Addition", "Skill B": "1D+1D sum upto 20",
                "# measurements where edge passes": 2,
                # Telangana columns present but should NOT be read.
                "Telangana Exit→Entry: P(B|A) — pooled": 0.95,
                "Telangana Exit→Entry: P(B|not A) — pooled": 0.60,
                # Delhi columns are the source of truth.
                "Delhi Entry→Entry: P(B|A) — pooled": 0.94,
                "Delhi Entry→Entry: P(B|not A) — pooled": 0.62,
            },
        ])
        edges = load_lattice_edges(str(path))
        assert len(edges) == 1
        e = edges[0]
        assert e.skill_a == "2-digit Addition with carry"
        assert e.skill_b == "1D+1D sum upto 20"
        assert e.p_b_given_a == 0.94    # Delhi, not Telangana 0.95
        assert e.p_b_given_not_a == 0.62  # Delhi, not Telangana 0.60
        assert e.weight == 1.0  # 2 measurements -> multi-view

    def test_telangana_only_edges_are_rejected(self, tmp_path: Path):
        """An edge with no Delhi values must raise, even if Telangana is populated.

        The engine is calibrated on Delhi data; we don't silently substitute
        Telangana values, since Telangana Exit→Entry and Delhi Entry→Entry
        measure different signals on different cohorts.
        """
        path = tmp_path / "lat.xlsx"
        _write_lattice_xlsx(path, rows=[
            {
                "Operation A": "Division", "Skill A": "X",
                "Operation B": "Multiplication", "Skill B": "Y",
                "# measurements where edge passes": 1,
                "Telangana Exit→Entry: P(B|A) — pooled": 0.92,
                "Telangana Exit→Entry: P(B|not A) — pooled": 0.55,
                "Delhi Entry→Entry: P(B|A) — pooled": None,
                "Delhi Entry→Entry: P(B|not A) — pooled": None,
            },
        ])
        with pytest.raises(CliInputError, match="Delhi"):
            load_lattice_edges(str(path))

    def test_delhi_missing_one_value_raises(self, tmp_path: Path):
        """Both Delhi columns must be present; partial coverage is an error."""
        path = tmp_path / "lat.xlsx"
        _write_lattice_xlsx(path, rows=[
            {
                "Operation A": "Addition", "Skill A": "X",
                "Operation B": "Addition", "Skill B": "Y",
                "# measurements where edge passes": 1,
                "Telangana Exit→Entry: P(B|A) — pooled": 0.9,
                "Telangana Exit→Entry: P(B|not A) — pooled": 0.5,
                "Delhi Entry→Entry: P(B|A) — pooled": 0.92,
                "Delhi Entry→Entry: P(B|not A) — pooled": None,  # missing
            },
        ])
        with pytest.raises(CliInputError, match="Delhi"):
            load_lattice_edges(str(path))

    def test_against_real_project_lattice(self):
        """Sanity-check against the real lattice_edges_final.xlsx in the repo data dir."""
        path = str(DATA_DIR / "lattice_edges_final.xlsx")
        if not Path(path).exists():
            pytest.skip("real lattice file not available in this environment")
        edges = load_lattice_edges(path)
        # Project records: 12 edges, 5 multi-view (weight 1.0), 7 single-view (weight 0.5).
        # The real file has Delhi Entry→Entry coverage for all 12 rows, so the
        # Delhi-only loader picks up the full set.
        assert len(edges) == 12
        multi = sum(1 for e in edges if e.weight >= 1.0)
        single = sum(1 for e in edges if e.weight < 1.0)
        assert multi == 5
        assert single == 7


# === cross_validate ========================================================


class TestCrossValidate:
    def test_no_warnings_when_clean(self):
        skills = [
            {"name": "A", "operation": "Addition", "sequence": 1, "content_grade": 1},
            {"name": "B", "operation": "Addition", "sequence": 2, "content_grade": 2},
        ]
        anchors = {3: {"Addition": "A"}}
        priors = {3: {"A": 0.8, "B": 0.5}}
        assert cross_validate(skills=skills, anchors=anchors, priors=priors) == []

    def test_unknown_anchor_warns(self):
        skills = [{"name": "A", "operation": "Addition", "sequence": 1, "content_grade": 1}]
        anchors = {3: {"Addition": "Not In Mapping"}}
        warnings = cross_validate(skills=skills, anchors=anchors, priors={})
        assert any("Not In Mapping" in w for w in warnings)

    def test_unknown_prior_skill_warns(self):
        skills = [{"name": "A", "operation": "Addition", "sequence": 1, "content_grade": 1}]
        priors = {3: {"Ghost Skill": 0.5}}
        warnings = cross_validate(skills=skills, anchors={}, priors=priors)
        assert any("Ghost Skill" in w for w in warnings)
